import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2,sheet.max_row+1): # need to add 1 because it doesn't include sheet.max_row
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4) # reference cell object in column 4
        corrected_price_cell.value = corrected_price

    # add chart
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)